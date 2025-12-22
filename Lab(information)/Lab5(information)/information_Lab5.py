# -*- coding: utf-8 -*-
# 实验室作业5 - 最简单版本 / Лабораторная работа 5 - упрощенная версия
# 学号: 408086 / Номер зачетной книжки: 408086

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os

print("=== Lab5 ===")

# 1. 文件路径 / Пути к файлам
input_file = r"D:\WST\python\Lab(information)\Lab5(information)\information-Lab5.xlsx"
output_file = r"D:\WST\python\Lab(information)\Lab5(information)\information-Lab5доп.xlsx"

print(f"Обработка файла: {os.path.basename(input_file)}")

# 2. 检查文件是否存在 / Проверка существования файла
if not os.path.exists(input_file):
    print(f"❌ Ошибка: файл не найден {input_file}")

    # 显示当前目录的文件 / Отображение файлов в текущем каталоге
    current_dir = os.path.dirname(input_file)
    print(f"\nФайлы в текущем каталоге {current_dir}:")
    try:
        for f in os.listdir(current_dir):
            if f.endswith(('.xlsx', '.xls', '.xlsm')):
                print(f"  - {f}")
    except:
        pass

    input("Нажмите Enter для выхода...")
    exit()

# 3. 用 pandas 读取Excel / Чтение Excel с помощью pandas
try:
    # 读取整个Excel / Чтение всего файла Excel
    df = pd.read_excel(input_file, header=None, engine='openpyxl')

    print(f"Исходный размер таблицы: {df.shape[0]} строк × {df.shape[1]} столбцов")

    # 选择 A4:Y15 区域 / Выбор области A4:Y15
    # 注意：pandas 索引从0开始 / Внимание: индексация pandas начинается с 0
    # A4 = 行3,列0 / A4 = строка 3, столбец 0
    # Y15 = 行14,列24 / Y15 = строка 14, столбец 24
    data = df.iloc[3:15, 0:25].copy()  # 行3-14,列0-24 / строки 3-14, столбцы 0-24

    print(f"✅ Успешно прочитана область A4:Y15")
    print(f"Размер данных: {data.shape[0]} строк × {data.shape[1]} столбцов")

    # 显示一些数据（可选，用于调试） / Отображение некоторых данных (опционально, для отладки)
    print("\nПредпросмотр первых строк данных:")
    print(data.head(3))

except Exception as e:
    print(f"❌ Ошибка чтения файла: {e}")
    input("Нажмите Enter для выхода...")
    exit()

# 4. 清空 F10:F15（F列是第6列，索引5） / Очистка области F10:F15 (столбец F - 6-й столбец, индекс 5)
if data.shape[1] >= 6:  # 确保有F列 / Убедиться, что есть столбец F
    print(f"\nОбработка столбца F (6-й столбец)...")

    # F10:F15 对应行索引9-14（在原始表格中是第10-15行）
    # F10:F15 соответствует индексам строк 9-14 (в исходной таблице строки 10-15)
    # 但在我们提取的A4:Y15中，索引是6-11
    # Но в извлеченной нами области A4:Y15 индексы 6-11
    for row in range(6, 12):  # F10-F15
        if row < len(data):
            data.iat[row, 5] = None  # 第6列（索引5）是F列 / 6-й столбец (индекс 5) - это столбец F
    print("✅ Область F10:F15 очищена")
else:
    print(f"⚠️  В таблице только {data.shape[1]} столбцов, нет столбца F")

# 5. 保存到新Excel文件 / Сохранение в новый файл Excel
try:
    data.to_excel(output_file, index=False, header=False, engine='openpyxl')
    print(f"\n✅ Данные сохранены в: {os.path.basename(output_file)}")
except Exception as e:
    print(f"❌ Ошибка сохранения: {e}")
    input("Нажмите Enter для выхода...")
    exit()

# 6. 设置边框 / Настройка границ
try:
    wb = load_workbook(output_file)
    ws = wb.active

    # 定义边框样式 / Определение стилей границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )

    # 给所有单元格加边框 / Добавление границ всем ячейкам
    cell_count = 0
    for row in range(1, data.shape[0] + 1):
        for col in range(1, data.shape[1] + 1):
            ws.cell(row=row, column=col).border = thin_border
            cell_count += 1

    print(f"✅ Границы добавлены для {cell_count} ячеек")

    # B1:C2 不加边框（B=2, C=3） / B1:C2 без границ (B=2, C=3)
    print("\nОсобая обработка области B1:C2...")
    for row in [1, 2]:      # 第1行和第2行 / строки 1 и 2
        for col in [2, 3]:  # B列和C列 / столбцы B и C
            ws.cell(row=row, column=col).border = no_border

    # 保存 / Сохранение
    wb.save(output_file)
    print("✅ Настройка границ завершена")

except Exception as e:
    print(f"❌ Ошибка настройки границ: {e}")
    input("Нажмите Enter для выхода...")
    exit()

# 7. 完成 / Завершение
print("\n" + "=" * 60)
print("🎉 Лабораторная работа №5 завершена!")
print("=" * 60)
print(f"📄 Исходный файл: {os.path.basename(input_file)}")
print(f"📄 Выходной файл: {os.path.basename(output_file)}")
print(f"📂 Расположение файлов: {os.path.dirname(output_file)}")
print("\n✅ Выполненные задачи:")
print("  1. Чтение области A4:Y15")
print("  2. Очистка области F10:F15")
print("  3. Добавление границ ячейкам")
print("  4. Область B1:C2 без границ")
print("=" * 60)

# 检查输出文件 / Проверка выходного файла
if os.path.exists(output_file):
    file_size = os.path.getsize(output_file) / 1024
    print(f"📊 Размер выходного файла: {file_size:.1f} KB")

    # 尝试打开结果文件 / Попытка открыть результирующий файл
    try:
        os.startfile(output_file)
        print("📂 Результирующий файл открыт")
    except:
        pass
else:
    print("⚠️  Предупреждение: выходной файл не найден")

input("\nНажмите Enter для выхода из программы...")